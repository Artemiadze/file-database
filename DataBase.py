import os
import pandas as pd
import openpyxl
from tkinter import messagebox


def create_empty_excel(columns: list, filename: str, sheet_name: str = 'Sheet1'):
    """ Функция создания базы данных в формате файлов Excel.
    Используется XlsxWriter в качестве движка для записи данных в файл Excel."""

    df = pd.DataFrame(columns=columns)  # Создание двумерной табличной структуры данных

    if not os.path.exists('Databases'):
        os.makedirs('Databases')

    filepath = os.path.join('Databases', filename)
    # используется XlsxWriter в качестве движка для записи данных в файл Excel.
    excel_writer = pd.ExcelWriter(filepath, engine='xlsxwriter')
    df.to_excel(excel_writer, index=False, sheet_name=sheet_name, freeze_panes=(1, 0))
    excel_writer._save()  # Сохранения изменений в файле Excel

    # return filepath


def parse_excel_to_dict_list(filepath: str, sheet_name='Sheet1'):
    """Запись данных из таблицы в словари для вывода данных на экран"""

    # Загружаем Excel файл в DataFrame
    df = pd.read_excel(filepath, sheet_name=sheet_name)

    # Преобразовывает каждую строку DataFrame в словарь, в котором ключами служат названия столбцов.
    dict_list = df.to_dict(orient='records')

    return dict_list


def get_data_to_ecxel(parametr, colomn ="nothing"):
    """ Функция выводит данные из словаря, по ключу, сам словарь создан в функции parse_excel_to_dict_list"""

    info = parse_excel_to_dict_list('Databases/student.xlsx')
    print_GUI = ""
    # Заголовок для вывода только нужного элемента, где строка с совпавшим ключом
    if colomn != "nothing":
        print_GUI = colomn.upper()

    for i in info:
        if parametr in i.values():
            if colomn == "nothing":
                print_GUI = print_GUI + "\n" + str(i)
            else:
                print_GUI = print_GUI + "\n" + str(i[colomn])
    messagebox.showinfo('Извлечение', print_GUI)


def insert_excel(information):
    """ Добавление в БД новую строку целиком, если её еще не было"""
    DB = openpyxl.load_workbook('Databases/student.xlsx')
    sheet = DB['Sheet1']

    info = parse_excel_to_dict_list('Databases/student.xlsx')
    flag = 0
    for i in info:
        if information[0] in i.values():
            print("There is already a field with such a key value!")
            flag = 1
    if flag == 0:
        sheet.append(information)
    DB.save('Databases/student.xlsx')


def delete_rows(number_deleted_rows):
    """ Удаление из БД строки по её номеру"""
    DB = openpyxl.load_workbook('Databases/student.xlsx')
    sheet = DB['Sheet1']
    sheet.delete_rows(number_deleted_rows)
    DB.save('Databases/student.xlsx')


def clean_database():
    """Очистка всей БД"""
    DB = openpyxl.load_workbook('Databases/student.xlsx')
    sheet = DB['Sheet1']
    sheet.delete_rows(2, sheet.max_row - 1)
    DB.save('Databases/student.xlsx')


def print_database():
    """ Вывод всей БД"""

    DB = openpyxl.load_workbook('Databases/student.xlsx')
    sheet = DB['Sheet1']
    print_to_console = ""
    for row in sheet.rows:
        string = ''
        for cell in row:
            string = string + str(cell.value) + ' '
        print_to_console = print_to_console + "\n" + string
    messagebox.showinfo('Таблица', print_to_console)
    DB.save('Databases/student.xlsx')


def print_backup():
    """Печать всей БД"""
    DB = openpyxl.load_workbook('Databases/backup.xlsx')
    sheet = DB['Sheet1']
    print_to_console = ""
    for row in sheet.rows:
        string = ''
        for cell in row:
            string = string + str(cell.value) + ' '
        print_to_console = print_to_console + "\n" + string
    messagebox.showinfo('Таблица', print_to_console)
    DB.save('Databases/student.xlsx')


def create_backup():
    """ Создаёт Backup файл и сохраняет туда данные из нашей базы данных"""
    # Копируются данные из нашей БД
    data = pd.read_excel('Databases/student.xlsx', sheet_name="Sheet1")

    create_empty_excel(columns=['ID', 'Name', 'Email', 'Group'], filename='backup.xlsx')
    # Заполнение Backup файла нашими данными из BD
    data.to_excel('Databases/backup.xlsx', sheet_name='Sheet1')

    # удаление первого столбца ,в котором перечисляются индексы
    myFile = openpyxl.load_workbook('Databases/student.xlsx')
    sheet_myFile = myFile['Sheet1']
    sheet_myFile.delete_cols(1)
    myFile.save('Databases/student.xlsx')

    backupFile = openpyxl.load_workbook('Databases/backup.xlsx')
    sheet_backup = backupFile['Sheet1']
    sheet_backup.delete_cols(1)
    backupFile.save('Databases/backup.xlsx')



